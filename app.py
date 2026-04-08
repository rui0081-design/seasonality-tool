import io
import re
import time
import random
import logging
import warnings
from pathlib import Path
from typing import Dict, Tuple, List

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from statsmodels.tsa.seasonal import seasonal_decompose
import plotly.graph_objects as go
import streamlit as st
from pytrends.request import TrendReq
from pytrends import exceptions as pytrends_exceptions
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.getLogger("matplotlib.font_manager").setLevel(logging.ERROR)
plt.rcParams["axes.unicode_minus"] = False
plt.rcParams["figure.dpi"] = 130
plt.rcParams["savefig.dpi"] = 180

# Japanese font fallback for matplotlib PNG export
try:
    import japanize_matplotlib  # noqa: F401
except Exception:
    pass

st.set_page_config(
    page_title="Google Trends Seasonality Analyzer",
    page_icon="📈",
    layout="wide",
)

GEO_MAP = {
    "日本 (JP)": "JP",
    "全世界": "",
    "アメリカ (US)": "US",
    "韓国 (KR)": "KR",
    "台湾 (TW)": "TW",
    "イギリス (GB)": "GB",
}

TIMEFRAME_MAP = {
    "過去12か月": "today 12-m",
    "過去3年": "today 3-y",
    "過去5年": "today 5-y",
    "2004年以降すべて": "all",
}

GPROP_MAP = {
    "通常のウェブ検索": "",
    "画像検索": "images",
    "ニュース検索": "news",
    "YouTube検索": "youtube",
    "Googleショッピング": "froogle",
}

MONTH_JP = {1: "1月", 2: "2月", 3: "3月", 4: "4月", 5: "5月", 6: "6月", 7: "7月", 8: "8月", 9: "9月", 10: "10月", 11: "11月", 12: "12月"}

st.markdown(
    """
    <style>
      :root {
        --bg:#FFFFFF;
        --panel:#F8FAFC;
        --line:#E2E8F0;
        --text:#0F172A;
        --sub:#475569;
        --blue:#1D4ED8;
        --navy:#0F172A;
        --amber:#D97706;
        --green:#059669;
        --red:#DC2626;
      }
      .block-container {max-width: 1200px; padding-top: 1.2rem; padding-bottom: 2rem;}
      h1,h2,h3 {color: var(--text);}
      .hero {
        padding: 18px 0 10px 0;
        margin-bottom: 10px;
      }
      .hero-title {font-size: 2.0rem; font-weight: 800; color: var(--navy); margin-bottom: 6px;}
      .hero-sub {font-size: 1rem; color: var(--sub); line-height: 1.8;}
      .hero-kicker {display:inline-block; font-size: .82rem; color: var(--blue); font-weight: 800; letter-spacing: .04em; margin-bottom: 8px;}
      .input-shell {
        border: 1px solid var(--line);
        border-radius: 20px;
        background: #FFFFFF;
        padding: 18px 18px 10px 18px;
        box-shadow: 0 10px 24px rgba(15,23,42,.04);
      }
      .section-label {font-size: .82rem; font-weight: 800; color: var(--blue); letter-spacing: .04em; margin-bottom: 4px;}
      .minimal-note {color: var(--sub); font-size: .92rem; line-height: 1.7;}
      .summary-grid {
        display:grid;
        grid-template-columns: repeat(4, minmax(0,1fr));
        gap: 12px;
        margin: 10px 0 18px 0;
      }
      .metric-card {
        border: 1px solid var(--line);
        border-radius: 18px;
        background: #FFFFFF;
        padding: 14px 16px;
      }
      .metric-label {font-size: .8rem; color: var(--sub); font-weight: 700; margin-bottom: 7px;}
      .metric-value {font-size: 1.45rem; color: var(--text); font-weight: 800; margin-bottom: 4px;}
      .metric-sub {font-size: .88rem; color: var(--sub); line-height: 1.6;}
      .insight-box {
        border-top: 1px solid var(--line);
        padding-top: 16px;
        margin-top: 10px;
      }
      .insight-item {
        border-left: 3px solid #DBEAFE;
        padding: 6px 0 6px 12px;
        margin-bottom: 10px;
      }
      .stDownloadButton button, .stButton button {
        border-radius: 14px; height: 2.8rem; font-weight: 700;
      }
      .stTabs [data-baseweb="tab-list"] {
        gap: 14px;
      }
      .stTabs [data-baseweb="tab"] {
        height: 42px; border-radius: 12px; padding-left: 16px; padding-right: 16px;
      }
      .dataframe tbody tr th:only-of-type {vertical-align: middle;}
      .dataframe tbody tr th {vertical-align: top;}
      .dataframe thead th {text-align: left;}
      @media (max-width: 900px) {
        .summary-grid {grid-template-columns: 1fr 1fr;}
      }
    </style>
    """,
    unsafe_allow_html=True,
)


def clean_name(text: str) -> str:
    text = re.sub(r'[\\/:*?"<>|]+', '_', str(text)).strip()
    return text[:80] if text else "google_trends_seasonality"


@st.cache_data(show_spinner=False, ttl=1800)
def fetch_google_trends_cached(keyword: str, geo: str, timeframe: str, gprop: str, cat: int, max_retries: int) -> pd.DataFrame:
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            pytrends = TrendReq(
                hl="ja-JP",
                tz=540,
                timeout=(10, 25),
                requests_args={
                    "headers": {
                        "User-Agent": (
                            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                            "AppleWebKit/537.36 (KHTML, like Gecko) "
                            "Chrome/124.0.0.0 Safari/537.36"
                        )
                    }
                },
            )
            time.sleep(random.uniform(0.8, 1.6))
            pytrends.build_payload([keyword], cat=int(cat), timeframe=timeframe, geo=geo, gprop=gprop)
            df = pytrends.interest_over_time()
            if df is None or df.empty:
                raise ValueError("Google Trendsからデータを取得できませんでした。")
            if "isPartial" in df.columns:
                df = df.drop(columns=["isPartial"])
            return df.rename(columns={keyword: "value"}).reset_index().rename(columns={"date": "date"})
        except Exception as e:
            last_err = e
            msg = str(e)
            too_many = ("429" in msg) or ("TooManyRequests" in msg) or isinstance(e, getattr(pytrends_exceptions, "TooManyRequestsError", tuple()))
            if attempt < max_retries:
                wait = 8 + (attempt - 1) * 8 if too_many else 3 + attempt * 2
                time.sleep(wait)
            else:
                raise last_err
    raise last_err


def infer_csv_columns(df: pd.DataFrame) -> Tuple[str, str]:
    date_candidates: List[str] = []
    value_candidates: List[str] = []
    for col in df.columns:
        if pd.to_datetime(df[col], errors="coerce").notna().mean() >= 0.6:
            date_candidates.append(col)
        if pd.to_numeric(df[col], errors="coerce").notna().mean() >= 0.6:
            value_candidates.append(col)
    if not date_candidates:
        raise ValueError("CSVから日付列を自動判定できませんでした。")
    if not value_candidates:
        raise ValueError("CSVから数値列を自動判定できませんでした。")
    return date_candidates[0], value_candidates[0]


def read_uploaded_csv(uploaded_file) -> pd.DataFrame:
    raw = uploaded_file.getvalue()
    for enc in ["utf-8-sig", "cp932", "utf-8", "shift_jis"]:
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=enc)
        except Exception:
            continue
    raise ValueError("CSVを読み込めませんでした。文字コードを見直してください。")


def normalize_monthly_index(df: pd.DataFrame, date_col: str, value_col: str) -> pd.DataFrame:
    ts = df[[date_col, value_col]].copy()
    ts.columns = ["date", "value"]
    ts["date"] = pd.to_datetime(ts["date"], errors="coerce")
    ts["value"] = pd.to_numeric(ts["value"], errors="coerce")
    ts = ts.dropna().sort_values("date")
    if ts.empty:
        raise ValueError("有効な日付列と数値列が見つかりませんでした。")
    ts = ts.set_index("date")

    inferred = pd.infer_freq(ts.index)
    if inferred and inferred.upper().startswith("W"):
        monthly = ts.resample("MS").mean()
    elif inferred and inferred.upper().startswith(("D", "B")):
        monthly = ts.resample("MS").mean()
    else:
        monthly = ts.resample("MS").mean()

    monthly["value"] = monthly["value"].interpolate(limit_direction="both")
    monthly = monthly[monthly.index.notna()]
    if monthly.shape[0] < 24:
        raise ValueError("季節性分析には最低24か月分の月次データが必要です。")
    return monthly


def safe_seasonal_decompose(series: pd.Series, period: int = 12):
    work = series.astype(float).copy()
    work = work.replace(0, np.nan).interpolate(limit_direction="both")
    work = work.fillna(method="bfill").fillna(method="ffill")
    if (work <= 0).any():
        work = work - work.min() + 1
    return seasonal_decompose(work, model="multiplicative", period=period, extrapolate_trend="freq")


def build_analysis_tables(ts: pd.DataFrame, keyword: str) -> Dict[str, pd.DataFrame]:
    result = ts.copy()
    decomp = safe_seasonal_decompose(result["value"], period=12)

    result["trend"] = decomp.trend
    result["seasonal"] = decomp.seasonal
    result["residual"] = decomp.resid
    result["adjusted"] = result["value"] / result["seasonal"]
    result["yoy"] = result["value"].pct_change(12)
    result["adjusted_yoy"] = result["adjusted"].pct_change(12)
    result["mom"] = result["value"].pct_change(1)
    result["adjusted_mom"] = result["adjusted"].pct_change(1)
    result["month_num"] = result.index.month
    result["month"] = result["month_num"].map(MONTH_JP)

    month_tbl = (
        result.groupby(["month_num", "month"], as_index=False)["seasonal"]
        .mean()
        .sort_values("month_num")
        .rename(columns={"seasonal": "seasonal_index"})
    )
    month_tbl["avg_diff"] = month_tbl["seasonal_index"] - 1
    month_tbl["avg_diff_pct"] = month_tbl["avg_diff"] * 100
    month_tbl["rank"] = month_tbl["seasonal_index"].rank(method="min", ascending=False).astype(int)

    latest = result.iloc[-1]
    prev12 = result.iloc[-13] if result.shape[0] >= 13 else None
    top3 = month_tbl.sort_values("seasonal_index", ascending=False).head(3)
    low3 = month_tbl.sort_values("seasonal_index", ascending=True).head(3)

    summary_df = pd.DataFrame([
        ["keyword", keyword],
        ["period", f"{result.index.min().strftime('%Y-%m')} ～ {result.index.max().strftime('%Y-%m')}"],
        ["latest_month", result.index.max().strftime('%Y-%m')],
        ["latest_value", round(float(latest['value']), 1)],
        ["latest_adjusted", round(float(latest['adjusted']), 1)],
        ["latest_yoy", None if pd.isna(latest['yoy']) else round(float(latest['yoy']) * 100, 1)],
        ["latest_adjusted_yoy", None if pd.isna(latest['adjusted_yoy']) else round(float(latest['adjusted_yoy']) * 100, 1)],
        ["strongest_month", top3.iloc[0]['month']],
        ["weakest_month", low3.iloc[0]['month']],
        ["seasonality_range", round(float(month_tbl['seasonal_index'].max() - month_tbl['seasonal_index'].min()), 3)],
    ], columns=["metric", "value"])

    recent_tbl = result[["value", "adjusted", "yoy", "adjusted_yoy", "trend", "seasonal"]].copy().tail(12).reset_index().rename(columns={"index": "date"})
    recent_tbl["date"] = recent_tbl["date"].dt.strftime("%Y-%m")
    recent_tbl["yoy"] = recent_tbl["yoy"] * 100
    recent_tbl["adjusted_yoy"] = recent_tbl["adjusted_yoy"] * 100

    detail = result.reset_index().rename(columns={"index": "date"})
    detail["date"] = pd.to_datetime(detail["date"]).dt.strftime("%Y-%m-%d")
    detail["yoy"] = detail["yoy"] * 100
    detail["adjusted_yoy"] = detail["adjusted_yoy"] * 100
    detail["mom"] = detail["mom"] * 100
    detail["adjusted_mom"] = detail["adjusted_mom"] * 100

    return {
        "summary": summary_df,
        "seasonality_by_month": month_tbl,
        "recent_result": recent_tbl,
        "result_detail": detail,
        "result_timeseries": result,
        "strong_top3": top3,
        "weak_top3": low3,
    }


def build_narrative(analysis: Dict[str, pd.DataFrame]) -> Dict[str, str]:
    month_tbl = analysis["seasonality_by_month"]
    result = analysis["result_timeseries"]
    top3 = analysis["strong_top3"]
    low3 = analysis["weak_top3"]
    latest = result.iloc[-1]

    strong_months = "・".join(top3["month"].tolist())
    weak_months = "・".join(low3["month"].tolist())
    adj_yoy = latest["adjusted_yoy"]
    raw_yoy = latest["yoy"]
    seasonal_idx = latest["seasonal"]
    seasonality_range_pct = (month_tbl["seasonal_index"].max() - month_tbl["seasonal_index"].min()) * 100

    if pd.isna(adj_yoy):
        momentum = "直近の前年比は算出条件を満たしておらず、伸びの判定は保留です。"
    elif adj_yoy >= 0.08:
        momentum = f"季節調整後前年比は {adj_yoy:.1%} で、直近の実力は明確に伸びています。"
    elif adj_yoy >= 0.02:
        momentum = f"季節調整後前年比は {adj_yoy:.1%} で、直近の実力は緩やかに上向きです。"
    elif adj_yoy > -0.02:
        momentum = f"季節調整後前年比は {adj_yoy:.1%} で、直近の実力はおおむね横ばいです。"
    else:
        momentum = f"季節調整後前年比は {adj_yoy:.1%} で、直近の実力は弱含みです。"

    if pd.isna(raw_yoy) or pd.isna(adj_yoy):
        driver = "季節要因と実力の切り分けは限定的です。"
    else:
        gap = raw_yoy - adj_yoy
        if seasonal_idx >= 1.05 and gap >= 0.05:
            driver = "足元は季節追い風の影響が強めです。見かけの伸びをそのまま実力と見ない方が安全です。"
        elif seasonal_idx <= 0.95 and gap <= -0.05:
            driver = "足元は季節逆風の影響が強めです。元データよりも実力は悪くない可能性があります。"
        elif abs(gap) <= 0.03:
            driver = "足元の動きは季節要因よりも実力変化の影響が中心です。"
        else:
            driver = "季節要因と実力要因の両方が動いています。片方だけでは説明しきれません。"

    if seasonality_range_pct >= 25:
        implication = f"季節差が大きい市場です。強い月（{strong_months}）の1〜2か月前に認知施策を厚くし、弱い月（{weak_months}）は効率重視で刈り取りに寄せる設計が有効です。"
    elif seasonality_range_pct >= 12:
        implication = f"季節性は中程度です。強い月（{strong_months}）の前倒し設計を基本にしつつ、弱い月（{weak_months}）は配信最適化やクリエイティブ検証に使うのが合理的です。"
    else:
        implication = f"季節性は強すぎません。年間で平準的に投資しつつ、直近の実力改善が見える局面に寄せてプロモーションを強化するのが妥当です。"

    return {
        "strong_vs_weak": f"需要が強い月は {strong_months}、弱い月は {weak_months} です。ピーク月とボトム月の差は約 {seasonality_range_pct:.1f}pt あります。",
        "momentum": momentum,
        "driver": driver,
        "implication": implication,
    }


def plotly_trend_chart(result: pd.DataFrame, keyword: str) -> go.Figure:
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=result.index,
        y=result["value"],
        mode="lines",
        name="実績",
        line=dict(width=2.8, color="#1D4ED8"),
        hovertemplate="%{x|%Y-%m}<br>実績: %{y:.1f}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=result.index,
        y=result["adjusted"],
        mode="lines",
        name="季節調整後",
        line=dict(width=2.8, color="#D97706"),
        hovertemplate="%{x|%Y-%m}<br>季節調整後: %{y:.1f}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=result.index,
        y=result["trend"],
        mode="lines",
        name="長期トレンド",
        line=dict(width=2.2, color="#64748B", dash="dot"),
        hovertemplate="%{x|%Y-%m}<br>トレンド: %{y:.1f}<extra></extra>",
    ))
    fig.update_layout(
        title=f"{keyword}｜検索需要の推移",
        height=440,
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        margin=dict(l=10, r=10, t=60, b=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
        xaxis=dict(showgrid=False, title=None),
        yaxis=dict(showgrid=True, gridcolor="#E2E8F0", title="検索指数"),
        hovermode="x unified",
    )
    return fig


def plotly_seasonality_chart(month_tbl: pd.DataFrame, keyword: str) -> go.Figure:
    bars_color = ["#1D4ED8" if x >= 1 else "#CBD5E1" for x in month_tbl["seasonal_index"]]
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=month_tbl["month"],
        y=month_tbl["seasonal_index"],
        marker_color=bars_color,
        text=[f"{v:.2f}" for v in month_tbl["seasonal_index"]],
        textposition="outside",
        hovertemplate="%{x}<br>季節指数: %{y:.3f}<extra></extra>",
        name="季節指数",
    ))
    fig.add_hline(y=1, line_dash="dash", line_color="#DC2626", opacity=0.75)
    fig.update_layout(
        title=f"{keyword}｜月別の季節指数",
        height=440,
        plot_bgcolor="#FFFFFF",
        paper_bgcolor="#FFFFFF",
        margin=dict(l=10, r=10, t=60, b=10),
        xaxis=dict(showgrid=False, title=None),
        yaxis=dict(showgrid=True, gridcolor="#E2E8F0", title="季節指数"),
        showlegend=False,
    )
    return fig


def create_dashboard_png(result: pd.DataFrame, month_tbl: pd.DataFrame, keyword: str) -> bytes:
    fig, axes = plt.subplots(2, 1, figsize=(12.5, 9.2), gridspec_kw={"height_ratios": [1.2, 1]})

    axes[0].plot(result.index, result["value"], linewidth=2.6, color="#1D4ED8", label="実績")
    axes[0].plot(result.index, result["adjusted"], linewidth=2.4, color="#D97706", label="季節調整後")
    axes[0].plot(result.index, result["trend"], linewidth=2.0, color="#64748B", linestyle="--", label="長期トレンド")
    axes[0].set_title(f"{keyword}｜検索需要の推移", loc="left", fontsize=16, fontweight="bold")
    axes[0].grid(axis="y", linestyle="--", alpha=0.22)
    axes[0].legend(frameon=False, ncol=3, loc="upper left")
    axes[0].spines[["top", "right"]].set_visible(False)
    axes[0].xaxis.set_major_locator(mdates.YearLocator())
    axes[0].xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    axes[0].set_ylabel("検索指数")

    vals = month_tbl["seasonal_index"].values
    colors = ["#1D4ED8" if v >= 1 else "#CBD5E1" for v in vals]
    bars = axes[1].bar(month_tbl["month"], vals, color=colors, width=0.68)
    axes[1].axhline(1, color="#DC2626", linestyle="--", linewidth=1.3, alpha=0.8)
    axes[1].set_title(f"{keyword}｜月別の季節指数", loc="left", fontsize=16, fontweight="bold")
    axes[1].grid(axis="y", linestyle="--", alpha=0.22)
    axes[1].spines[["top", "right"]].set_visible(False)
    axes[1].set_ylabel("季節指数")
    axes[1].set_ylim(max(0, vals.min() - 0.12), vals.max() + 0.12)
    for bar, v in zip(bars, vals):
        axes[1].text(bar.get_x() + bar.get_width()/2, v + 0.012, f"{v:.2f}", ha="center", va="bottom", fontsize=9)

    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def metric_guide_df() -> pd.DataFrame:
    return pd.DataFrame([
        ["value", "Google Trendsの元データ。大きいほど検索需要が強い。"],
        ["trend", "短期変動をならした長期的な基調。"],
        ["seasonal / seasonal_index", "月ごとの季節性。1.00が平均、1.10なら平均より約10%強い。"],
        ["residual", "トレンドと季節性で説明しきれない一時的な変動。"],
        ["adjusted", "季節要因を除いた実力ベースの値。"],
        ["yoy", "元データの前年同月比（%）。"],
        ["adjusted_yoy", "季節要因を除いた前年同月比（%）。"],
        ["mom / adjusted_mom", "前月比（%）。月次の変化を見る補助指標。"],
    ], columns=["metric", "description"])


def how_to_read_df() -> pd.DataFrame:
    return pd.DataFrame([
        ["まず見る", "strongest_month / weakest_month", "いつ需要が上がりやすく、落ちやすい市場かを把握する。"],
        ["次に見る", "latest_adjusted_yoy", "季節要因を除いた実力の伸びを確認する。"],
        ["見分け方", "latest_yoy と latest_adjusted_yoy の差", "差が大きいほど、見かけの伸び縮みは季節要因の影響が大きい。"],
        ["施策化", "強い月の1〜2か月前", "認知施策や販促の仕込み時期として使う。"],
        ["運用", "弱い月", "効率重視の配信、クリエイティブ検証、指名刈り取りに向く。"],
    ], columns=["step", "watch_metric", "how_to_use"])


def make_excel_bytes(analysis: Dict[str, pd.DataFrame], narrative: Dict[str, str], keyword: str) -> bytes:
    summary_export = analysis["summary"].copy()
    summary_export = pd.concat([
        summary_export,
        pd.DataFrame([
            ["summary_strong_vs_weak", narrative["strong_vs_weak"]],
            ["summary_momentum", narrative["momentum"]],
            ["summary_driver", narrative["driver"]],
            ["summary_implication", narrative["implication"]],
        ], columns=["metric", "value"])
    ], ignore_index=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_export.to_excel(writer, sheet_name="summary", index=False)
        analysis["seasonality_by_month"].to_excel(writer, sheet_name="seasonality_by_month", index=False)
        analysis["result_detail"].to_excel(writer, sheet_name="result_detail", index=False)
        metric_guide_df().to_excel(writer, sheet_name="metric_guide", index=False)
        how_to_read_df().to_excel(writer, sheet_name="how_to_read", index=False)

    output.seek(0)
    wb = load_workbook(output)

    header_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
    header_font = Font(name="Yu Gothic UI", size=10, bold=True, color="0F172A")
    body_font = Font(name="Yu Gothic UI", size=10, color="0F172A")
    border = Border(
        left=Side(style="thin", color="D9E2F1"),
        right=Side(style="thin", color="D9E2F1"),
        top=Side(style="thin", color="D9E2F1"),
        bottom=Side(style="thin", color="D9E2F1"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.auto_filter.ref = ws.dimensions
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
            width = min(max(max_len + 2, 12), 44)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.sheet_view.showGridLines = False

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def result_csv_bytes(detail_df: pd.DataFrame) -> bytes:
    return detail_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def friendly_error_message(err: Exception) -> str:
    msg = str(err)
    if "429" in msg or "TooManyRequests" in msg:
        return "Google Trends側が一時的に混み合っています。少し待って再実行するか、期間を短くして試してください。"
    if "24か月" in msg:
        return "季節性分析には最低24か月分の月次データが必要です。期間を長くしてください。"
    return f"処理中にエラーが発生しました。入力条件かデータ形式を見直してください。詳細: {msg}"


# Header
st.markdown(
    """
    <div class="hero">
      <div class="hero-kicker">GOOGLE TRENDS / SEASONALITY ANALYZER</div>
      <div class="hero-title">検索需要のトレンド・季節性・直近の実力を、一画面で読む</div>
      <div class="hero-sub">
        Keywordを入れるだけで、Google Trendsから検索需要を取得し、季節性分解・月別季節指数・直近の実力評価まで整理します。<br>
        社内共有向けのExcel、CSV、PNGもそのまま出力できます。
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

left, right = st.columns([1.8, 1], gap="large")

with left:
    st.markdown('<div class="input-shell">', unsafe_allow_html=True)
    st.markdown('<div class="section-label">INPUT</div>', unsafe_allow_html=True)
    with st.form("main_form", clear_on_submit=False):
        keyword = st.text_input(
            "Keyword",
            placeholder="例：NISA、花粉、転職、air max",
            help="まずは単語1つでの分析がおすすめです。",
        )
        c1, c2, c3 = st.columns(3)
        with c1:
            geo_label = st.selectbox("地域", list(GEO_MAP.keys()), index=0)
        with c2:
            timeframe_label = st.selectbox("期間", list(TIMEFRAME_MAP.keys()), index=2)
        with c3:
            source_mode = st.selectbox("データ取得", ["Google Trends", "CSVアップロード"], index=0)

        uploaded_file = None
        gprop_label = "通常のウェブ検索"
        cat = 0
        retry_count = 5

        if source_mode == "CSVアップロード":
            uploaded_file = st.file_uploader("CSVをアップロード", type=["csv"])
            st.caption("日付列と数値列が入ったCSVを読み込みます。週次データは月次に集約します。")
        else:
            with st.expander("詳細設定", expanded=False):
                d1, d2, d3 = st.columns(3)
                with d1:
                    gprop_label = st.selectbox("検索種別", list(GPROP_MAP.keys()), index=0)
                with d2:
                    cat = st.number_input("カテゴリID", min_value=0, step=1, value=0)
                with d3:
                    retry_count = st.slider("再試行回数", min_value=2, max_value=6, value=5)
                st.caption("通常はデフォルトのままで問題ありません。")

        submitted = st.form_submit_button("分析開始", type="primary", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

with right:
    st.markdown("<div class='section-label'>HOW TO USE</div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div class="minimal-note">
        1. Keyword / 地域 / 期間を入力<br>
        2. 分析開始を押す<br>
        3. 需要が強い月・弱い月・直近の実力を確認<br>
        4. 必要ならExcel / CSV / PNGでそのまま共有
        </div>
        """,
        unsafe_allow_html=True,
    )
    with st.expander("指標の見方", expanded=False):
        st.dataframe(metric_guide_df(), use_container_width=True, hide_index=True)

if submitted:
    try:
        keyword_clean = keyword.strip()
        if source_mode == "Google Trends":
            if not keyword_clean:
                st.warning("Keywordを入力してください。")
                st.stop()
            with st.spinner("Google Trendsから取得して分析しています…"):
                raw_df = fetch_google_trends_cached(
                    keyword=keyword_clean,
                    geo=GEO_MAP[geo_label],
                    timeframe=TIMEFRAME_MAP[timeframe_label],
                    gprop=GPROP_MAP[gprop_label],
                    cat=cat,
                    max_retries=retry_count,
                )
                ts = normalize_monthly_index(raw_df, raw_df.columns[0], "value")
        else:
            if uploaded_file is None:
                st.warning("CSVファイルをアップロードしてください。")
                st.stop()
            with st.spinner("CSVを読み込んで分析しています…"):
                raw_csv = read_uploaded_csv(uploaded_file)
                date_col, value_col = infer_csv_columns(raw_csv)
                ts = normalize_monthly_index(raw_csv, date_col, value_col)
                if not keyword_clean:
                    keyword_clean = Path(uploaded_file.name).stem

        analysis = build_analysis_tables(ts, keyword_clean)
        narrative = build_narrative(analysis)
        result = analysis["result_timeseries"]
        month_tbl = analysis["seasonality_by_month"]

        latest = result.iloc[-1]
        strongest = analysis["strong_top3"].iloc[0]["month"]
        weakest = analysis["weak_top3"].iloc[0]["month"]
        adj_yoy_val = latest["adjusted_yoy"]
        adj_yoy_txt = "-" if pd.isna(adj_yoy_val) else f"{adj_yoy_val:.1%}"
        season_range = month_tbl["seasonal_index"].max() - month_tbl["seasonal_index"].min()

        st.markdown("<div class='section-label' style='margin-top:18px;'>SUMMARY</div>", unsafe_allow_html=True)
        st.markdown(
            f"""
            <div class="summary-grid">
              <div class="metric-card">
                <div class="metric-label">需要が強い月</div>
                <div class="metric-value">{strongest}</div>
                <div class="metric-sub">上位3か月: {'・'.join(analysis['strong_top3']['month'].tolist())}</div>
              </div>
              <div class="metric-card">
                <div class="metric-label">需要が弱い月</div>
                <div class="metric-value">{weakest}</div>
                <div class="metric-sub">下位3か月: {'・'.join(analysis['weak_top3']['month'].tolist())}</div>
              </div>
              <div class="metric-card">
                <div class="metric-label">直近の実力</div>
                <div class="metric-value">{adj_yoy_txt}</div>
                <div class="metric-sub">季節調整後前年比</div>
              </div>
              <div class="metric-card">
                <div class="metric-label">季節差の大きさ</div>
                <div class="metric-value">{season_range*100:.1f}pt</div>
                <div class="metric-sub">ピーク月とボトム月の差</div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown("<div class='insight-box'>", unsafe_allow_html=True)
        st.markdown(f"<div class='insight-item'>{narrative['strong_vs_weak']}</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='insight-item'>{narrative['momentum']}</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='insight-item'>{narrative['driver']}</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='insight-item'>{narrative['implication']}</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        trend_fig = plotly_trend_chart(result, keyword_clean)
        season_fig = plotly_seasonality_chart(month_tbl, keyword_clean)

        tab1, tab2, tab3 = st.tabs(["グラフ", "テーブル", "ダウンロード"])

        with tab1:
            g1, g2 = st.columns(2)
            with g1:
                st.plotly_chart(trend_fig, use_container_width=True, config={"displayModeBar": False})
            with g2:
                st.plotly_chart(season_fig, use_container_width=True, config={"displayModeBar": False})

        with tab2:
            t1, t2 = st.columns([1, 1], gap="large")
            with t1:
                st.subheader("月別季節指数")
                display_month = month_tbl.copy()
                display_month["seasonal_index"] = display_month["seasonal_index"].round(3)
                display_month["avg_diff_pct"] = display_month["avg_diff_pct"].round(1)
                st.dataframe(display_month[["rank", "month", "seasonal_index", "avg_diff_pct"]], use_container_width=True, hide_index=True)
            with t2:
                st.subheader("直近結果")
                display_recent = analysis["recent_result"].copy()
                for col in ["value", "adjusted", "trend", "seasonal"]:
                    display_recent[col] = display_recent[col].round(1)
                for col in ["yoy", "adjusted_yoy"]:
                    display_recent[col] = display_recent[col].round(1)
                st.dataframe(display_recent.tail(6), use_container_width=True, hide_index=True)

        with tab3:
            excel_data = make_excel_bytes(analysis, narrative, keyword_clean)
            csv_data = result_csv_bytes(analysis["result_detail"])
            png_data = create_dashboard_png(result, month_tbl, keyword_clean)
            safe = clean_name(keyword_clean)
            d1, d2, d3 = st.columns(3)
            with d1:
                st.download_button(
                    "Excelをダウンロード",
                    excel_data,
                    file_name=f"{safe}_seasonality_analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d2:
                st.download_button(
                    "CSVをダウンロード",
                    csv_data,
                    file_name=f"{safe}_result_detail.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with d3:
                st.download_button(
                    "PNGをダウンロード",
                    png_data,
                    file_name=f"{safe}_dashboard.png",
                    mime="image/png",
                    use_container_width=True,
                )
            st.caption("Excelには summary / seasonality_by_month / result_detail / metric_guide / how_to_read の5シートが入ります。")

    except Exception as e:
        st.error(friendly_error_message(e))
        with st.expander("見直しポイント"):
            st.markdown(
                "- Google Trendsなら期間を長めにするか、少し具体的なKeywordにする\n"
                "- 同じ条件で短時間に連続実行しすぎない\n"
                "- CSVなら日付列と数値列が入っているか確認する\n"
                "- 季節性分析には24か月以上の月次データが必要"
            )
